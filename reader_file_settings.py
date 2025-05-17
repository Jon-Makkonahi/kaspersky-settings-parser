"""Файл для хранения функций для чтения эксель файлов"""
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from instrumentarium import dump_json
from forming_dictionary import forming_dictionary


def find_last_non_empty_value(sheet: Worksheet, row: int, column: int):
    """Функция находит последнее значение != NoneType в строке выше"""
    for row_low in range(row, 3, -1):
        value = sheet.cell(row=row_low, column=column).value
        if value is not None:
            return value
    return None


def read_objects(workbook: Workbook, sheetname: str, flag=False):
    """Функция для чтения объектов из файла ksc_settings."""
    sheet = workbook[sheetname]
    max_col = sheet.max_column
    max_row = sheet.max_row
    objects_with_types: dict = {}
    count = 0
    # формируем ключи в словаре по шапке файла, лист "policies"
    for column in range(1, max_col + 1):
        name_header = sheet.cell(row=1, column=column).value
        objects_with_types.setdefault(name_header, [])
    # заполняем значениями словарь, лист "policies"
    for row in range(2, max_row + 1):
        count += 1
        for column in range(1, max_col + 1):
            name_header = sheet.cell(row=1, column=column).value
            value = sheet.cell(row=row, column=column).value
            objects_with_types[name_header].append(value)
    objects_with_types['Настройки'] = [''] * count
    if flag:
        dump_json('objects.json', objects_with_types)
    return objects_with_types


def read_settings(workbook: Workbook, check_sheet_name: str):
    """Функция для чтения настроек из файла ksc_settings."""
    sheet = workbook[check_sheet_name]
    settings_for_type: dict = {}
    max_col = sheet.max_column
    max_row = sheet.max_row
    value = None
    # формируем ключи в словаре по шапке файла, лист "название типа объекта"
    for column in range(1, max_col + 1):
        name_header = sheet.cell(row=3, column=column).value
        settings_for_type.setdefault(name_header, [])
    # заполняем значениями словарь, лист "название типа объекта"
    for row in range(4, max_row + 1):
        # в словарь вносят проверки параметров значения,
        # которые включены в "Аудит"
        if sheet.cell(row=row, column=6).value == 'ON':
            for column in range(1, max_col + 1):
                name_header = sheet.cell(row=3, column=column).value
                value = sheet.cell(row=row, column=column).value
                # замена пустых значений NoneType на корректные
                # исключение только для 1 и 2 столбца
                if value is None and column not in [1, 2]:
                    value = find_last_non_empty_value(sheet, row, column)
                settings_for_type[name_header].append(value)
    return settings_for_type


def all_settings_and_object_with_type(sheetname: str, flag=False) -> tuple:
    """Функция управления остальными функциями."""
    all_settings: dict[str, dict] = {}
    workbook = load_workbook('ksc_settings.xlsx')
    objects_with_type = read_objects(workbook, sheetname, flag=flag)
    if objects_with_type['Индекс']:
        for idx, name_type in enumerate(objects_with_type['Тип проверки']):
            name_type = name_type[
                :name_type.find('_')] if '_' in name_type else name_type
            check_sheet_name = objects_with_type['Чек-лист'][idx]
            settings_for_type = read_settings(workbook, check_sheet_name)
            # в словарь помещаются все настройки с режимом "Аудит"
            # согласно типу политики
            all_settings[check_sheet_name] = settings_for_type
            # данная функция формирует словари для проверки
            # такие словари заносят в переменную checkers файла checkers.py
            forming_dictionary(check_sheet_name, all_settings, flag=False)
        workbook.close()
        if flag:
            dump_json('settings.json', all_settings)
        return all_settings, objects_with_type
    return (False, False)


if __name__ == '__main__':
    all_settings_and_object_with_type('profiles', flag=True)
