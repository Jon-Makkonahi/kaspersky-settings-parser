"""Файл, в котором размещены функции записывающие результаты"""
from typing import cast
from openpyxl import Workbook

from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Стили
THIN_BORDER = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)
ORANGE = PatternFill(
    fill_type='solid', start_color='ffa500', end_color='ffa500'
)
YELLOW = PatternFill(
    fill_type='solid', start_color='ffff00', end_color='ffff00'
)
RED = PatternFill(fill_type='solid', start_color='ff0000', end_color='ff0000')
GREEN = PatternFill(
    fill_type='solid', start_color='9ACD32', end_color='9ACD32'
)
PURPLE = PatternFill(
    fill_type='solid', start_color='884488', end_color='884488'
)
FONT_HEADERS = Font(name='Times New Roman', bold=True, size=12)
FONT_BODY = Font(name='Times New Roman', size=12)


def write_result_policies(result: list[str], sheet: Worksheet):
    """Функция, записывает результат обработки политик в Excel"""

    headers = [
        'Название', 'Группа', 'Пункт стандарта', 'Требование стандарта',
        'Навигация по параметрам KSC', 'Настройка KSC',
        'Запрет на редактирование', 'Требуемое значение настройки',
        'Текущее значение запрета на редактирования',
        'Текущее значение настройки', 'Статус наблюдения'
    ]
    column_widths = [30, 25, 25, 25, 50, 40, 15, 15, 25, 25, 25]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 9:
                if columns[8] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[8] != 'Соответствует':
                    cell.fill = ORANGE
            elif col_idx == 10:
                if columns[9] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[9] != 'Соответствует':
                    cell.fill = YELLOW
            elif col_idx == 11:
                if columns[10] == 'NEW':
                    cell.fill = GREEN
    # Настройки листа
    sheet.freeze_panes = 'D2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_result_profiles(result: list[str], sheet: Worksheet):
    """Функция, записывает результат обработки профилей в Excel"""
    headers = [
        'Название политики', 'Группа', 'Профиль', 'Пункт стандарта',
        'Требование стандарта', 'Навигация по параметрам KSC', 'Настройка KSC',
        'Запрет на редактирование', 'Требуемое значение настройки',
        'Текущее значение запрета на редактирования',
        'Текущее значение настройки', 'Статус наблюдения'
    ]
    column_widths = [30, 25, 30, 25, 25, 50, 40, 15, 15, 25, 25, 25]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 10:
                if columns[9] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[9] != 'Соответствует':
                    cell.fill = ORANGE
            elif col_idx == 11:
                if columns[10] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[10] != 'Соответствует':
                    cell.fill = YELLOW
            elif col_idx == 12:
                if columns[11] == 'NEW':
                    cell.fill = GREEN
    # Настройки листа
    sheet.freeze_panes = 'E2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_result_tasks(result: list[str], sheet: Worksheet):
    """Функция, записывает результат обработки задач в Excel"""

    headers = [
        'Название', 'Группа', 'Пункт стандарта',
        'Требование стандарта', 'Навигация по параметрам KSC', 'Настройка KSC',
        'Требуемое значение настройки', 'Текущее значение настройки',
        'Статус наблюдения'
    ]
    column_widths = [30, 25, 25, 25, 50, 40, 15, 25, 25]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 8:
                if columns[7] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[7] != 'Соответствует':
                    cell.fill = YELLOW
            elif col_idx == 9:
                if columns[8] == 'NEW':
                    cell.fill = GREEN
    # Настройки листа
    sheet.freeze_panes = 'D2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_policies(result: list[str], sheet: Worksheet):
    """Функция, записывает политики в Excel"""

    headers = [
        'Индекс', 'Название', 'Группа', 'Тип программы', 'Наследование',
        'Статус', 'Статус наблюдения'
    ]
    column_widths = [12, 40, 35, 35, 15, 15, 12]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width
    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 7:
                if columns[6] == 'NEW':
                    cell.fill = GREEN
                elif columns[6] == 'DEL':
                    cell.fill = PURPLE
    # Настройки листа
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_profiles(result: list[str], sheet: Worksheet):
    """Функция, записывает политики в Excel"""

    headers = [
        'Индекс', 'Название политики', 'Группа', 'Профиль', 'Тип программы',
        'Статус', 'Статус наблюдения'
    ]
    column_widths = [12, 40, 25, 25, 25, 15, 12]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width
    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 7:
                if columns[6] == 'NEW':
                    cell.fill = GREEN
                elif columns[6] == 'DEL':
                    cell.fill = PURPLE
    # Настройки листа
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_tasks(result: list[str], sheet: Worksheet):
    """Функция, записывает политики в Excel"""

    headers = [
        'Индекс', 'Название', 'Группа', 'Тип программы', 'Тип задачи',
        'Статус наблюдения'
    ]
    column_widths = [12, 40, 25, 25, 25, 12]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width
    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 6:
                if columns[5] == 'NEW':
                    cell.fill = GREEN
                elif columns[5] == 'DEL':
                    cell.fill = PURPLE
    # Настройки листа
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_activate_policies(result: list[str], sheet: Worksheet):
    """Функция, записывает политики в Excel"""

    headers = [
        'Индекс', 'Название', 'Группа', 'Тип программы', 'Наследование',
        'Статус наблюдения'
    ]
    column_widths = [12, 40, 35, 35, 15, 15]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width
    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 6:
                if columns[5] == 'Изменение было':
                    cell.fill = PURPLE
    # Настройки листа
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_activate_profiles(result: list[str], sheet: Worksheet):
    """Функция, записывает политики в Excel"""

    headers = [
        'Индекс', 'Название политики', 'Группа', 'Профиль', 'Тип программы',
        'Статус наблюдения'
    ]
    column_widths = [12, 40, 35, 35, 15, 15]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width
    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 6:
                if columns[5] == 'Изменение было':
                    cell.fill = PURPLE
    # Настройки листа
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


WRITE_FUNCTION = {
    'policies_inconsistencies': write_result_policies,
    'tasks_inconsistencies': write_result_tasks,
    'profiles_inconsistencies': write_result_profiles,
    'policies': write_policies,
    'profiles': write_profiles,
    'tasks': write_tasks,
    'policies_activate': write_activate_policies,
    'profiles_activate': write_activate_profiles
}


def write_all(result: dict, filename: str):
    """Функция, записывает результат обработки в Excel"""
    wb = Workbook()
    default_sheet = cast(Worksheet, wb.active)
    for inconsistencies, nested_data in result.items():
        sheet = wb.create_sheet(title=nested_data['sheetname'])
        WRITE_FUNCTION[inconsistencies](nested_data['data'], sheet)
    wb.remove(default_sheet)
    # Сохранение
    wb.save(filename)
    wb.close()


def write_policies_inconsistencies(result: list[str], filename: str,
                                   fl_pf=False):
    """Функция, записывает результат обработки в Excel"""
    wb = Workbook()
    sheet = cast(Worksheet, wb.active)
    sheet.title = 'Отчет'

    headers = [
        'Пункт стандарта', 'Требование стандарта', 'Параметры',
        'Настройка', 'Запрет на редактирование',
        'Требуемое значение настройки',
        'Текущее значение запрета на редактирования',
        'Текущее значение настройки'
    ]
    column_widths = [25, 25, 50, 40, 15, 15, 25, 25]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for row_idx, line in enumerate(result, start=2):
        if fl_pf:
            columns = line.replace('None', '').split('|')[3:]
        else:
            columns = line.replace('None', '').split('|')[2:]
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 7:
                if columns[6] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[6] != 'Соответствует':
                    cell.fill = ORANGE
            elif col_idx == 8:
                if columns[7] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[7] != 'Соответствует':
                    cell.fill = YELLOW
    # Настройки листа
    sheet.freeze_panes = 'C2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Сохранение
    wb.save(filename)
    wb.close()


def write_tasks_inconsistencies(result: list[str], filename: str):
    """Функция, записывает результат обработки в Excel"""
    wb = Workbook()
    sheet = cast(Worksheet, wb.active)
    sheet.title = 'Отчет'

    headers = [
        'Пункт стандарта', 'Требование стандарта', 'Параметры',
        'Настройка', 'Требуемое значение настройки',
        'Текущее значение настройки'
    ]
    column_widths = [25, 25, 50, 40, 15, 25]
    # Заголовки
    for col, (title, width) in enumerate(zip(headers, column_widths), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.border = THIN_BORDER
        cell.font = FONT_HEADERS
        sheet.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for row_idx, line in enumerate(result, start=2):  # Начнем со 2-й строки
        columns = line.replace('None', '').split('|')[2:]
        for col_idx, value in enumerate(columns, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = FONT_BODY
            if col_idx == 6:
                if columns[5] == 'Нет значения в JSON':
                    cell.fill = RED
                elif columns[5] != 'Соответствует':
                    cell.fill = YELLOW
    # Настройки листа
    sheet.freeze_panes = 'C2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Сохранение
    wb.save(filename)
    wb.close()


if __name__ == '__main__':
    pass
